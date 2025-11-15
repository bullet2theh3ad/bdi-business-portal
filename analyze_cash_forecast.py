#!/usr/bin/env python3
"""
Analyze BDI Cash Forecast Excel file structure
Goal: Understand how to optimize tabs to roll up to CURRENT WEEK CASHFLOW FORECAST
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import sys

# Load the workbook
file_path = "25.11.10 BDI Cash Forecast.xlsx"
print(f"📊 Analyzing: {file_path}\n")
print("=" * 80)

try:
    # Load with openpyxl to get sheet info
    wb = load_workbook(file_path, data_only=True)
    
    print("\n📋 SHEET STRUCTURE:")
    print("=" * 80)
    for i, sheet_name in enumerate(wb.sheetnames, 1):
        sheet = wb[sheet_name]
        print(f"\n{i}. Sheet: '{sheet_name}'")
        print(f"   Dimensions: {sheet.dimensions}")
        print(f"   Max Row: {sheet.max_row}, Max Col: {sheet.max_column}")
        
        # Try to read first few rows to understand structure
        print(f"   First 5 rows preview:")
        for row_idx in range(1, min(6, sheet.max_row + 1)):
            row_data = []
            for col_idx in range(1, min(6, sheet.max_column + 1)):
                cell = sheet.cell(row_idx, col_idx)
                val = str(cell.value)[:30] if cell.value else ""
                row_data.append(val)
            print(f"      Row {row_idx}: {' | '.join(row_data)}")
    
    print("\n" + "=" * 80)
    print("\n📊 DETAILED ANALYSIS OF KEY SHEETS:")
    print("=" * 80)
    
    # Analyze CURRENT WEEK CASHFLOW FORECAST
    if "CURRENT WEEK CASHFLOW FORECAST" in wb.sheetnames:
        print("\n🎯 CURRENT WEEK CASHFLOW FORECAST (Main Tab)")
        print("-" * 80)
        sheet = wb["CURRENT WEEK CASHFLOW FORECAST"]
        
        # Find date headers (usually in first few rows)
        print("\n   Looking for date/week structure...")
        for row_idx in range(1, min(15, sheet.max_row + 1)):
            row_vals = [sheet.cell(row_idx, col).value for col in range(1, min(20, sheet.max_column + 1))]
            # Check if row contains dates or week indicators
            if any(val and ("Week" in str(val) or "2024" in str(val) or "2025" in str(val)) for val in row_vals):
                print(f"\n   Row {row_idx} (potential date/week headers):")
                for col_idx, val in enumerate(row_vals[:15], 1):
                    if val:
                        print(f"      Col {col_idx}: {val}")
        
        # Find category labels (left column)
        print("\n   Category structure (Column A):")
        for row_idx in range(1, min(50, sheet.max_row + 1)):
            cell_val = sheet.cell(row_idx, 1).value
            if cell_val and len(str(cell_val)) > 2:
                print(f"      Row {row_idx}: {cell_val}")
    
    # Analyze Weekly Spend
    if "Weekly Spend" in wb.sheetnames:
        print("\n\n💰 WEEKLY SPEND (Rollup Tab)")
        print("-" * 80)
        sheet = wb["Weekly Spend"]
        
        print("\n   Structure preview (first 20 rows, first 10 cols):")
        for row_idx in range(1, min(21, sheet.max_row + 1)):
            row_data = []
            for col_idx in range(1, min(11, sheet.max_column + 1)):
                cell = sheet.cell(row_idx, col_idx)
                val = str(cell.value)[:25] if cell.value else ""
                row_data.append(val)
            print(f"      Row {row_idx}: {' | '.join(row_data)}")
    
    # Analyze Labor Spend
    if "Labor Spend" in wb.sheetnames:
        print("\n\n👥 LABOR SPEND (Detail Tab)")
        print("-" * 80)
        sheet = wb["Labor Spend"]
        
        print("\n   Current structure (first 25 rows, first 8 cols):")
        for row_idx in range(1, min(26, sheet.max_row + 1)):
            row_data = []
            for col_idx in range(1, min(9, sheet.max_column + 1)):
                cell = sheet.cell(row_idx, col_idx)
                val = str(cell.value)[:30] if cell.value else ""
                row_data.append(val)
            print(f"      Row {row_idx}: {' | '.join(row_data)}")
        
        print("\n   ⚠️  ISSUES TO FIX:")
        print("      - Need proper categories: Operations, Marketing, G&A, R&D")
        print("      - Need employee/contractor breakdown")
        print("      - Need pay frequency tracking")
        print("      - Need department mapping")
    
    # Look for other expense detail tabs
    print("\n\n📑 OTHER DETAIL TABS:")
    print("-" * 80)
    detail_tabs = [s for s in wb.sheetnames if s not in ["CURRENT WEEK CASHFLOW FORECAST", "Weekly Spend"]]
    for sheet_name in detail_tabs:
        sheet = wb[sheet_name]
        print(f"\n   {sheet_name}:")
        print(f"      Rows: {sheet.max_row}, Cols: {sheet.max_column}")
        
        # Check if it has formulas linking to Weekly Spend
        has_formulas = False
        for row in sheet.iter_rows(min_row=1, max_row=min(20, sheet.max_row)):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    has_formulas = True
                    break
            if has_formulas:
                break
        print(f"      Contains formulas: {has_formulas}")
    
    print("\n" + "=" * 80)
    print("\n🔍 OPTIMIZATION RECOMMENDATIONS:")
    print("=" * 80)
    
    print("\n1. DYNAMIC DATE RANGE:")
    print("   - Replace hardcoded week columns with formula-based dates")
    print("   - Use TODAY() function to anchor current week")
    print("   - Extend 13 weeks backward and 13 weeks forward automatically")
    print("   - Highlight current week with conditional formatting")
    
    print("\n2. LABOR SPEND RESTRUCTURE:")
    print("   - Add columns: Employee Name, Department, Category, Pay Frequency")
    print("   - Categories: Operations, Marketing, G&A, R&D, Contract Labor")
    print("   - Pay Frequency: Weekly, Bi-Weekly, Monthly, Contract")
    print("   - Auto-calculate weekly amounts based on frequency")
    
    print("\n3. WEEKLY SPEND IMPROVEMENTS:")
    print("   - Add subtotals by category")
    print("   - Add variance column (Actual vs Forecast)")
    print("   - Add YTD tracking")
    print("   - Link to detail tabs with drill-down capability")
    
    print("\n4. CURRENT WEEK CASHFLOW FORECAST:")
    print("   - Add rolling 13-week view (dynamically updates)")
    print("   - Add cumulative cash position row")
    print("   - Add cash runway indicator")
    print("   - Add variance analysis section")
    
    print("\n5. OTHER EXPENSE CATEGORIES NEEDED:")
    print("   - Marketing Spend (broken down by channel)")
    print("   - OpEx by category (rent, insurance, subscriptions)")
    print("   - Inventory/COGS purchases")
    print("   - NRE/R&D spend")
    print("   - Debt service / financing costs")
    
    wb.close()
    
    print("\n" + "=" * 80)
    print("\n✅ Analysis complete!")
    print("\nNext: Export detailed structure to create redesign plan")
    
except FileNotFoundError:
    print(f"❌ Error: File '{file_path}' not found!")
    print("   Make sure the file is in the current directory.")
    sys.exit(1)
except Exception as e:
    print(f"❌ Error analyzing file: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)

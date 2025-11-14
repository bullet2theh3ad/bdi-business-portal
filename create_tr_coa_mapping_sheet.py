#!/usr/bin/env python3
"""
Create TR COA Mapping sheet and add to NRE DevOps Restructuring Plan
"""

import pandas as pd
import json
from openpyxl import load_workbook

# Read the TR COA file
tr_coa_file = "/Users/Steve/Projects/BDI/BDI PORTAL/TB TR COA Mapping baseline.xlsx"
nre_plan_file = "/Users/Steve/Projects/BDI/BDI PORTAL/NRE_DevOps_Restructuring_Plan.xlsx"

print("=" * 80)
print("CREATING TR COA MAPPING SHEET")
print("=" * 80)
print()

# Read TR COA Workday sheet
print("📖 Reading TR COA (Workday COA sheet)...")
df_tr = pd.read_excel(tr_coa_file, sheet_name='Workday COA', header=0)
print(f"✓ Loaded {len(df_tr)} TR COA accounts")
print()

# Clean up column names
df_tr.columns = ['Account_Description', 'Reporting_Template', 'TR_GL_Code', 'IS_BS', 'EBITDA_Mapping', 'Comments']

# Filter to only rows with GL codes
df_tr = df_tr[df_tr['TR_GL_Code'].notna()].copy()
df_tr = df_tr[df_tr['TR_GL_Code'] != 'GL Code'].copy()  # Remove header duplicates

print(f"✓ Found {len(df_tr)} TR GL Code accounts")
print()

# Create mapping for our new accounts to TR GL codes
mapping_data = []

# R&D (9000s) → TR GL Mapping
rd_mapping = [
    {'New_Account': 9000, 'New_Name': 'Research & Development', 'TR_GL_Code': '6000', 'TR_Description': 'R&D Expenses', 'Mapping_Logic': 'Parent account for all R&D expenses'},
    {'New_Account': 9010, 'New_Name': 'Product Engineering & Design', 'TR_GL_Code': '6010', 'TR_Description': 'Engineering', 'Mapping_Logic': 'Product development engineering'},
    {'New_Account': 9020, 'New_Name': 'Firmware Development', 'TR_GL_Code': '6010', 'TR_Description': 'Engineering', 'Mapping_Logic': 'Firmware = engineering labor'},
    {'New_Account': 9030, 'New_Name': 'R&D Documentation', 'TR_GL_Code': '6010', 'TR_Description': 'Engineering', 'Mapping_Logic': 'Technical documentation'},
    
    {'New_Account': 9100, 'New_Name': 'Product Compliance', 'TR_GL_Code': '6000', 'TR_Description': 'R&D Expenses', 'Mapping_Logic': 'Parent for compliance costs'},
    {'New_Account': 9110, 'New_Name': 'Product Certifications (UL, FCC, CE)', 'TR_GL_Code': '6050', 'TR_Description': 'Certifications & Testing', 'Mapping_Logic': 'Product certifications'},
    {'New_Account': 9120, 'New_Name': 'Compliance Testing & Validation', 'TR_GL_Code': '6050', 'TR_Description': 'Certifications & Testing', 'Mapping_Logic': 'Testing & validation'},
    {'New_Account': 9130, 'New_Name': 'Regulatory & Registration Fees', 'TR_GL_Code': '6050', 'TR_Description': 'Certifications & Testing', 'Mapping_Logic': 'Regulatory fees'},
    
    {'New_Account': 9200, 'New_Name': 'Prototyping & Tooling Expense', 'TR_GL_Code': '6040', 'TR_Description': 'Prototyping & Tooling (expensed)', 'Mapping_Logic': 'Parent for expensed tooling'},
    {'New_Account': 9210, 'New_Name': 'Mold Design & Engineering (Expensed)', 'TR_GL_Code': '6040', 'TR_Description': 'Prototyping & Tooling', 'Mapping_Logic': 'Low-value molds (<$2.5k)'},
    {'New_Account': 9220, 'New_Name': 'Prototype Materials & Services', 'TR_GL_Code': '6040', 'TR_Description': 'Prototyping & Tooling', 'Mapping_Logic': 'Prototype materials'},
    {'New_Account': 9230, 'New_Name': 'Test Fixtures & Jigs (Expensed)', 'TR_GL_Code': '6040', 'TR_Description': 'Prototyping & Tooling', 'Mapping_Logic': 'Low-value test equipment'},
    
    {'New_Account': 9300, 'New_Name': 'R&D Services', 'TR_GL_Code': '6000', 'TR_Description': 'R&D Expenses', 'Mapping_Logic': 'Parent for external R&D services'},
    {'New_Account': 9310, 'New_Name': 'Engineering Consultants', 'TR_GL_Code': '6020', 'TR_Description': 'Engineering Contractors', 'Mapping_Logic': 'External engineering consultants'},
    {'New_Account': 9320, 'New_Name': 'Technical Documentation Services', 'TR_GL_Code': '6020', 'TR_Description': 'Engineering Contractors', 'Mapping_Logic': 'External documentation services'},
    {'New_Account': 9330, 'New_Name': 'R&D Contract Labor', 'TR_GL_Code': '6020', 'TR_Description': 'Engineering Contractors', 'Mapping_Logic': '1099 contractors for R&D'},
]

# DevOps (6500s) → TR GL Mapping
devops_mapping = [
    {'New_Account': 6500, 'New_Name': 'DevOps & Infrastructure', 'TR_GL_Code': '5300', 'TR_Description': 'IT & Technology', 'Mapping_Logic': 'Parent for DevOps/IT operational costs'},
    {'New_Account': 6510, 'New_Name': 'Cloud Hosting Services', 'TR_GL_Code': '5310', 'TR_Description': 'Cloud Services', 'Mapping_Logic': 'AWS, Azure, cloud hosting'},
    {'New_Account': 6511, 'New_Name': 'AWS Services', 'TR_GL_Code': '5310', 'TR_Description': 'Cloud Services', 'Mapping_Logic': 'Amazon Web Services'},
    {'New_Account': 6512, 'New_Name': 'Azure Services', 'TR_GL_Code': '5310', 'TR_Description': 'Cloud Services', 'Mapping_Logic': 'Microsoft Azure'},
    {'New_Account': 6513, 'New_Name': 'Other Cloud Providers', 'TR_GL_Code': '5310', 'TR_Description': 'Cloud Services', 'Mapping_Logic': 'Google Cloud, other providers'},
    
    {'New_Account': 6520, 'New_Name': 'Infrastructure Support', 'TR_GL_Code': '5320', 'TR_Description': 'IT Support & Monitoring', 'Mapping_Logic': 'Infrastructure management tools'},
    {'New_Account': 6521, 'New_Name': 'Server Monitoring & Management', 'TR_GL_Code': '5320', 'TR_Description': 'IT Support & Monitoring', 'Mapping_Logic': 'Datadog, New Relic, etc.'},
    {'New_Account': 6522, 'New_Name': 'CI/CD Pipeline Tools', 'TR_GL_Code': '5320', 'TR_Description': 'IT Support & Monitoring', 'Mapping_Logic': 'Jenkins, GitHub Actions'},
    {'New_Account': 6523, 'New_Name': 'Infrastructure Automation', 'TR_GL_Code': '5320', 'TR_Description': 'IT Support & Monitoring', 'Mapping_Logic': 'Terraform, Ansible'},
    
    {'New_Account': 6530, 'New_Name': 'Application Hosting', 'TR_GL_Code': '5310', 'TR_Description': 'Cloud Services', 'Mapping_Logic': 'Database, CDN, load balancers'},
    {'New_Account': 6531, 'New_Name': 'Database Hosting', 'TR_GL_Code': '5310', 'TR_Description': 'Cloud Services', 'Mapping_Logic': 'RDS, managed databases'},
    {'New_Account': 6532, 'New_Name': 'CDN & Load Balancing', 'TR_GL_Code': '5310', 'TR_Description': 'Cloud Services', 'Mapping_Logic': 'CloudFront, Cloudflare'},
    
    {'New_Account': 6540, 'New_Name': 'DevOps Tools & Licenses', 'TR_GL_Code': '5330', 'TR_Description': 'Software & Subscriptions', 'Mapping_Logic': 'DevOps software licenses'},
    {'New_Account': 6541, 'New_Name': 'Development Tools & IDEs', 'TR_GL_Code': '5330', 'TR_Description': 'Software & Subscriptions', 'Mapping_Logic': 'Visual Studio, IntelliJ'},
    {'New_Account': 6542, 'New_Name': 'DevOps Platforms', 'TR_GL_Code': '5330', 'TR_Description': 'Software & Subscriptions', 'Mapping_Logic': 'GitHub, GitLab subscriptions'},
]

# Labor (7000s) → TR GL Mapping
labor_mapping = [
    {'New_Account': 7000, 'New_Name': 'Labor Expenses', 'TR_GL_Code': '4000', 'TR_Description': 'Payroll & Benefits', 'Mapping_Logic': 'Parent for all labor costs'},
    {'New_Account': 7210, 'New_Name': 'Engineering Contractors', 'TR_GL_Code': '4100', 'TR_Description': 'Contract Labor', 'Mapping_Logic': 'Engineering 1099 contractors (split R&D vs G&A)'},
    {'New_Account': 7220, 'New_Name': 'Operations Contractors', 'TR_GL_Code': '4100', 'TR_Description': 'Contract Labor', 'Mapping_Logic': 'Operations 1099 contractors'},
    {'New_Account': 7230, 'New_Name': 'Marketing Contractors', 'TR_GL_Code': '4100', 'TR_Description': 'Contract Labor', 'Mapping_Logic': 'Marketing 1099 contractors'},
]

# Depreciation (9600s) → TR GL Mapping
depreciation_mapping = [
    {'New_Account': 9600, 'New_Name': 'Depreciation & Amortization', 'TR_GL_Code': '7000', 'TR_Description': 'Depreciation & Amortization', 'Mapping_Logic': 'Parent for all D&A'},
    {'New_Account': 9610, 'New_Name': 'Depreciation - Tooling & Molds', 'TR_GL_Code': '7010', 'TR_Description': 'Depreciation - Tooling', 'Mapping_Logic': 'Depreciation of capitalized tooling (FA 1510)'},
    {'New_Account': 9620, 'New_Name': 'Depreciation - Manufacturing Equipment', 'TR_GL_Code': '7020', 'TR_Description': 'Depreciation - Equipment', 'Mapping_Logic': 'Depreciation of production equipment (FA 1520)'},
    {'New_Account': 9630, 'New_Name': 'Depreciation - Test Equipment', 'TR_GL_Code': '7020', 'TR_Description': 'Depreciation - Equipment', 'Mapping_Logic': 'Depreciation of test equipment (FA 1530)'},
    {'New_Account': 9640, 'New_Name': 'Depreciation - IT Equipment', 'TR_GL_Code': '7030', 'TR_Description': 'Depreciation - IT Hardware', 'Mapping_Logic': 'Depreciation of computers, servers (FA 1530)'},
    {'New_Account': 9650, 'New_Name': 'Amortization - Software', 'TR_GL_Code': '7040', 'TR_Description': 'Amortization - Software', 'Mapping_Logic': 'Amortization of capitalized software (FA 1540)'},
]

# Payroll Liabilities (2300s) → TR GL Mapping
payroll_mapping = [
    {'New_Account': 2310, 'New_Name': 'Salaries & Wages Payable', 'TR_GL_Code': '2100', 'TR_Description': 'Payroll Liabilities', 'Mapping_Logic': 'Accrued unpaid wages'},
    {'New_Account': 2326, 'New_Name': 'Federal Income Tax Withheld', 'TR_GL_Code': '2110', 'TR_Description': 'Tax Withholdings', 'Mapping_Logic': 'Federal tax withheld (trust fund)'},
    {'New_Account': 2327, 'New_Name': 'State Income Tax Withheld', 'TR_GL_Code': '2110', 'TR_Description': 'Tax Withholdings', 'Mapping_Logic': 'State tax withheld (trust fund)'},
    {'New_Account': 2328, 'New_Name': 'FICA/Medicare - Employee Portion', 'TR_GL_Code': '2110', 'TR_Description': 'Tax Withholdings', 'Mapping_Logic': 'FICA withheld (trust fund)'},
    {'New_Account': 2329, 'New_Name': 'Employee 401(k) Deferrals', 'TR_GL_Code': '2120', 'TR_Description': '401k Deferrals', 'Mapping_Logic': '401k employee deferrals (ERISA)'},
    {'New_Account': 2350, 'New_Name': '401(k) Employer Match Payable', 'TR_GL_Code': '2120', 'TR_Description': '401k Deferrals', 'Mapping_Logic': 'Accrued employer 401k match'},
    {'New_Account': 2360, 'New_Name': 'Health Insurance - Employer Portion', 'TR_GL_Code': '2130', 'TR_Description': 'Benefits Payable', 'Mapping_Logic': 'Accrued health insurance premiums'},
    {'New_Account': 2370, 'New_Name': 'Workers Comp Insurance Payable', 'TR_GL_Code': '2130', 'TR_Description': 'Benefits Payable', 'Mapping_Logic': 'Accrued workers comp premiums'},
]

# Combine all mappings
mapping_data = rd_mapping + devops_mapping + labor_mapping + depreciation_mapping + payroll_mapping

print(f"✓ Created mappings for {len(mapping_data)} new accounts")
print()

# Convert to DataFrame
df_mapping = pd.DataFrame(mapping_data)

# Add category column
df_mapping['Category'] = ''
df_mapping.loc[df_mapping['New_Account'].between(9000, 9399), 'Category'] = 'R&D (NRE)'
df_mapping.loc[df_mapping['New_Account'].between(6500, 6599), 'Category'] = 'DevOps (OpEx)'
df_mapping.loc[df_mapping['New_Account'].between(7000, 7299), 'Category'] = 'Labor'
df_mapping.loc[df_mapping['New_Account'].between(9600, 9699), 'Category'] = 'Depreciation'
df_mapping.loc[df_mapping['New_Account'].between(2300, 2399), 'Category'] = 'Payroll Liabilities'

# Reorder columns
df_mapping = df_mapping[[
    'Category',
    'New_Account',
    'New_Name',
    'TR_GL_Code',
    'TR_Description',
    'Mapping_Logic'
]]

print("=" * 80)
print("ADDING SHEET TO NRE DEVOPS RESTRUCTURING PLAN")
print("=" * 80)
print()

# Load the existing workbook
try:
    # Use openpyxl directly
    book = load_workbook(nre_plan_file)
    
    # Check if sheet already exists and remove it
    if '10_TR_COA_Mapping' in book.sheetnames:
        del book['10_TR_COA_Mapping']
        print("✓ Removed existing TR COA Mapping sheet")
    
    # Create a new sheet using ExcelWriter
    with pd.ExcelWriter(nre_plan_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        df_mapping.to_excel(writer, sheet_name='10_TR_COA_Mapping', index=False)
    
    print("✓ Added Sheet 10: TR COA Mapping")
    print(f"✓ Mapped {len(df_mapping)} accounts to TR GL codes")
    print()
    
    # Print summary by category
    print("=" * 80)
    print("MAPPING SUMMARY BY CATEGORY:")
    print("=" * 80)
    for category in df_mapping['Category'].unique():
        count = len(df_mapping[df_mapping['Category'] == category])
        print(f"  {category}: {count} accounts")
    
    print()
    print("=" * 80)
    print("✅ TR COA MAPPING COMPLETE")
    print("=" * 80)
    print(f"📁 File: {nre_plan_file}")
    print()
    print("📊 New Sheet Added:")
    print("  • Sheet 10: TR COA Mapping (New Account → TR GL Code)")
    print()
    print("🎯 Use this mapping to:")
    print("  • Map BOSS portal GL codes to TR reporting codes")
    print("  • Ensure consistency between BOSS and TR reporting")
    print("  • Create automated GL code translation rules")
    print()
    print("=" * 80)
    
except Exception as e:
    print(f"❌ Error: {e}")
    import traceback
    traceback.print_exc()

